import warnings
warnings.filterwarnings("ignore")
from openpyxl import load_workbook
from openpyxl import Workbook

def call_excel(year):
    sum_men_die = 0
    sum_men = 0
    sum_women_die = 0
    sum_women = 0
    wb = load_workbook(year + '.xlsx')
    sh = wb[year]
    for i in range(7, 83):
        sum_men += sh['D'+str(i)].value # Find sum of men who live.
        sum_women += sh['E'+str(i)].value # Find sum of women who live.
        sum_men_die += sh['G'+str(i)].value # Find sum of men who dead form suicide.
        sum_women_die += sh['H'+str(i)].value # Find sum of women who dead form suicide.
    total_people = sum_men + sum_women
    total_die = sum_men_die + sum_women_die
    mean_men_die = sum_men_die * 1000000 / sum_men
    mean_women_die = sum_women_die * 1000000 / sum_women
    total_die = mean_men_die + mean_women_die
    return  mean_men_die, mean_women_die, total_die

def write_excel():
    wb = Workbook()
    sh = wb.active
    sh.merge_cells('B1:D1')
    sh['B1'] = "Number of suicides per million population."
    sh['B2'] = "Man"
    sh['A2'] = "Years(man)"
    sh['D2'] = "Woman"
    sh['C2'] = "Years(woman)"
    sh['F2'] = "Total"
    sh['E2'] = "Years(total)"
    years = ["2548", "2549", "2550", "2551", "2552", "2553", "2554", "2555", "2556", "2557"]
    collum = 3
    for year in years: # Append the result values to each cell.
        mean_men_die, mean_women_die, total_die = call_excel(year)
        sh['A'+str(collum)] = year
        sh['C'+str(collum)] = year
        sh['E'+str(collum)] = year
        sh['B'+str(collum)] = "%.2f" % mean_men_die
        sh['D'+str(collum)] = "%.2f" % mean_women_die
        sh['F'+str(collum)] = "%.2f" % total_die
        collum += 1
    wb.save('result.xlsx')
    result10_man, result10_woman, result10_total = predict_10years() # Append the predict values to result.xlsx .
    sh['A13'], sh['C13'], sh['E13'] = "2558(Predict)", "2558(Predict)", "2558(Predict)"
    sh['B13'], sh['D13'], sh['F13'] = result10_man, result10_woman, result10_total
    wb.save('result.xlsx')

def predict_10years(): # Calculate the predict values.
    wb = load_workbook('result.xlsx')
    sh = wb['Sheet']
    result10_man, result10_woman, result10_total = [], [], []
    for row in ["B", "D", "F"]: # Append the values to list.
        for column in range(3, 13):
                if row == "B":
                    result10_man.append(float(sh[row + str(column)].value))
                elif row == "D":
                    result10_woman.append(float(sh[row + str(column)].value))
                else: result10_total.append(float(sh[row + str(column)].value))
    return "%.2f" %(sum(result10_man) / 10), "%.2f" %(sum(result10_woman) / 10), \
             "%.2f" %(sum(result10_total) / 10)

write_excel()
##print(predict_10years())
