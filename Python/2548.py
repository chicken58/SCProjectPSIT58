import warnings
warnings.filterwarnings("ignore")
from openpyxl import load_workbook
from openpyxl import Workbook
def call_excel():
    sum_men_die = 0
    sum_men = 0
    sum_women_die = 0
    sum_women = 0
    wb = load_workbook('2548.xlsx')
    sh = wb['Sheet2']
    for i in range(7, 83):
        sum_men += sh['D'+str(i)].value
        sum_women += sh['E'+str(i)].value
        sum_men_die += sh['G'+str(i)].value
        sum_women_die += sh['H'+str(i)].value
    total_people = sum_men + sum_women
    total_die = sum_men_die + sum_women_die
    mean_men_die = sum_men_die * 1000000 / sum_men
    mean_women_die = sum_women_die * 1000000 / sum_women
    mean_total_die = total_die * 1000000 / total_people
    return  mean_men_die, mean_women_die, mean_total_die


def write_excel():
    mean_men_die, mean_women_die, mean_total_die = call_excel()
    wb = Workbook()
    sh = wb.active
    sh.merge_cells('B1:D1')
    sh['B1'] = "จำนวนคนฆ่าตัวตายต่อประชากรล้านคน"
    sh['B2'] = "ผู้ชาย"
    sh['A2'] = "ปี(ชาย)"
    sh['D2'] = "ผู้หญิง"
    sh['C2'] = "ปี(หญิง)"
    sh['F2'] = "รวม"
    sh['E2'] = "ปี(รวม)"
    sh['A3'] = "พ.ศ.2548"
    sh['C3'] = "พ.ศ.2548"
    sh['E3'] = "พ.ศ.2548"
    sh['B3'] = "%.2f" % mean_men_die
    sh['D3'] = "%.2f" % mean_women_die
    sh['F3'] = "%.2f" % mean_total_die
    wb.save('result.xlsx')

write_excel()
